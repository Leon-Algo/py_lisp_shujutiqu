import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def format_excel(file_path='devices_with_quantities.xlsx'):
    """
    格式化Excel文件，隐藏指定列
    """
    try:
        # 加载工作簿
        wb = load_workbook(file_path)
        
        # 处理"设备明细"sheet
        if '设备明细' in wb.sheetnames:
            ws = wb['设备明细']
            
            # 查找需要隐藏的列
            headers = [cell.value for cell in ws[1]]
            hide_columns = ['坐标', '数量文本坐标', '匹配距离']
            
            # 隐藏指定列
            for col_idx, header in enumerate(headers, 1):
                if header in hide_columns:
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].hidden = True
        
        # 保存更改
        wb.save(file_path)
        print(f"\nExcel格式化完成: {file_path}")
        print("已隐藏列: 坐标, 数量文本坐标, 匹配距离")
        return True
        
    except Exception as e:
        print(f"格式化Excel时出错: {str(e)}")
        return False 